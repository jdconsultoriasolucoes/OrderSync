import openpyxl

file_path = r"E:\OrderSync\backend\assets\template_supra.xlsx"
try:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    print("Sheets in template_supra.xlsx:", wb.sheetnames)
    if len(wb.sheetnames) > 1:
        ws2 = wb.worksheets[1]
        print(f"Analyzing second sheet: {ws2.title}")
        for row in ws2.iter_rows():
            for cell in row:
                if cell.value is not None:
                    print(f"{cell.coordinate}: {cell.value}")
    else:
        print("No second sheet in template_supra.")
except Exception as e:
    print(f"Error: {e}")
