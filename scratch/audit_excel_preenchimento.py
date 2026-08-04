import openpyxl
from pathlib import Path

template_path = Path(r'e:\OrderSync\backend\assets\template_supra.xlsx')
wb = openpyxl.load_workbook(template_path)

# Map of cells touched by excel_supra_service.py in Sheet 1:
sheet1_touched = [
    "A8", "A9", "A10", "A11", "E11", "A12", "A13", "E13",
    "E17", "C18", "E18", "G18", "G17", "I18", "I17",
    "A21", "I21", "A22", "F22", "I22", "A23",
    "A25", "I25", "A26", "F26", "I26",
    "A29", "F29", "A30", "F30",
    "A34..A37", "C34..C37", "E34..E37", # Ref bancarias
    "A40..A43", "E40..E43", "G40..G43", "I40..I43", # Ref comerciais
    "A46..A48", "C46..C48", "F46..F48", "H46..H48", "J46..J48", # Bens imoveis
    "A57..A59", "E57..E59", "G57..G59", "I57..I59", # Bens moveis
    "A51..A53", "F51..F53", "H51..H53", "J51..J53", # Planteis
    "C61" # Local e Data
]

# Map of cells touched by excel_supra_service.py in Sheet 2:
sheet2_touched = [
    "E7", "E8", "C14", "C15", "C16", "A9", "A10",
    "C20", "C21", "A22", "C22",
    "E25", "H25", "E26", "H26", "E28", "H28",
    "A41", "A42", "A43", "D34", "H34", "D35"
]

print("=== ABA 1: Cadastro Parte 1 ===")
ws1 = wb["Cadastro Parte 1"]
for r in range(1, ws1.max_row + 1):
    row_str = []
    for c in range(1, ws1.max_column + 1):
        cell = ws1.cell(r, c)
        val = str(cell.value or '').strip()
        if val:
            row_str.append(f"{cell.coordinate}: '{val}'")
    if row_str:
        print(f"Row {r:02d}: " + " | ".join(row_str))

print("\n=== ABA 2: Cadastro Parte 2 ===")
ws2 = wb["Cadastro Parte 2"]
for r in range(1, ws2.max_row + 1):
    row_str = []
    for c in range(1, ws2.max_column + 1):
        cell = ws2.cell(r, c)
        val = str(cell.value or '').strip()
        if val:
            row_str.append(f"{cell.coordinate}: '{val}'")
    if row_str:
        print(f"Row {r:02d}: " + " | ".join(row_str))
