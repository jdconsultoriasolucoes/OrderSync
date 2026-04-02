
import openpyxl
from reportlab.lib.units import cm

template_path = r"E:\Projeto Sistema pedidos\Planejamento\Arquivos bases edson\NOVA_FICHA_DE_CADASTRO_ALISUL (1).xlsx"
wb = openpyxl.load_workbook(template_path, data_only=True)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n=== Geometria da Aba: {sheet_name} ===")
    
    # Extrair larguras de colunas (estão em "unidades de caractere" no Excel, converter aprox)
    col_widths = []
    for i in range(1, ws.max_column + 1):
        letter = openpyxl.utils.get_column_letter(i)
        w = ws.column_dimensions[letter].width
        if w is None: w = 8.43 # largura padrão
        # Conversão aproximada: 1 unit = ~0.2cm (depende da fonte, mas usaremos como base)
        col_widths.append(w)
    print(f"Col Widths (Excel units): {col_widths}")
    
    # Extrair alturas de linhas
    row_heights = []
    for i in range(1, 40): # primeiras 40 linhas
        h = ws.row_dimensions[i].height
        if h is None: h = 15.0 # altura padrão
        row_heights.append(h)
    print(f"Row Heights (points): {row_heights}")
    
    # Detectar mesclagens (MergeCells)
    print("Células Mescladas:")
    for merge in ws.merged_cells.ranges:
        print(f" - {merge}")
