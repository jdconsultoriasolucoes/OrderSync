import io
import logging
from datetime import datetime
from pathlib import Path
import openpyxl

from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.lib import colors
from reportlab.pdfgen import canvas
from reportlab.platypus import Table, TableStyle

from .excel_supra_service import gerar_excel_cliente_supra

logger = logging.getLogger("ordersync.pdf_supra")

# Assets
_BASE = Path(__file__).resolve().parent.parent / "assets"
_LOGO_PATH  = _BASE / "img_aba_Cadastro_Parte_1_0.png"
_QR_PATH    = _BASE / "img_aba_Cadastro_Parte_1_1.png"

def _s(val):
    return str(val) if val is not None else ""

def gerar_pdf_cliente_supra(cli) -> bytes:
    """
    Gera o PDF do Cadastro Supra baseado DIRETAMENTE no preenchimento do Excel (Aba 1).
    Isso garante que o layout seja idêntico ao Excel.
    """
    def hex_to_color(hex_str):
        if not hex_str or hex_str == '00000000' or len(hex_str) < 6:
            return None
        if len(hex_str) == 8: # ARGB
            hex_str = hex_str[2:]
        try:
            r = int(hex_str[0:2], 16) / 255.0
            g = int(hex_str[2:4], 16) / 255.0
            b = int(hex_str[4:6], 16) / 255.0
            return colors.Color(r, g, b)
        except:
            return None

    try:
        # 1. Gera o Excel preenchido para servir de base
        excel_bytes = gerar_excel_cliente_supra(cli)
        wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), data_only=True)
        ws = wb["Cadastro Parte 1"]
        
        buffer = io.BytesIO()
        c = canvas.Canvas(buffer, pagesize=A4)
        w, h = A4
        margin = 0.4 * cm # Margem técnica mínima para impressão profissional

        # Área de Impressão Fixa (A1:K64)
        max_r = 64
        max_c = 11
        matrix = [["" for _ in range(max_c)] for _ in range(max_r)]
        
        # Configurações do Motor de Impressão
        style_list = [
            ('FONTNAME', (0,0), (-1,-1), 'Helvetica'),
            ('FONTSIZE', (0,0), (-1,-1), 6.5),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('LEFTPADDING', (0,0), (-1,-1), 1.8),
            ('RIGHTPADDING', (0,0), (-1,-1), 1.8),
            ('TOPPADDING', (0,0), (-1,-1), 1),
            ('BOTTOMPADDING', (0,0), (-1,-1), 1),
        ]

        # 2. Varredura dinâmica de células e estilos
        for r in range(1, max_r + 1):
            for c_idx in range(1, max_c + 1):
                cell = ws.cell(row=r, column=c_idx)
                
                # Valor
                if cell.value is not None:
                    matrix[r-1][c_idx-1] = _s(cell.value)
                
                coord = (c_idx-1, r-1)
                
                # Alinhamento Automático
                if cell.alignment:
                    ha = cell.alignment.horizontal
                    if ha in ['center', 'right']:
                        style_list.append(('ALIGN', coord, coord, ha.upper()))
                
                # Estilo de Fonte (Bold)
                if cell.font and cell.font.bold:
                    style_list.append(('FONTNAME', coord, coord, 'Helvetica-Bold'))
                
                # Tamanho da Fonte (Ajuste de Escala Excel -> PDF)
                if cell.font and cell.font.size:
                    fs = cell.font.size * 0.72 # Escala de pontos
                    style_list.append(('FONTSIZE', coord, coord, max(5, fs)))

        # 3. Geometria Dinâmica e Escala (Fit-to-Page)
        def get_col_width(idx):
            cw = ws.column_dimensions[openpyxl.utils.get_column_letter(idx+1)].width
            if not cw: cw = 10
            return cw * 5.4
            
        def get_row_height(idx):
            rh = ws.row_dimensions[idx+1].height
            if not rh: rh = 13.5
            return rh

        col_w_raw = [get_col_width(i) for i in range(max_c)]
        total_w_excel = sum(col_w_raw)
        
        # Fator de escala para largura da página A4
        scale_f = (w - 2*margin) / total_w_excel
        col_w = [cw * scale_f for cw in col_w_raw]
        row_h = [get_row_height(i) * scale_f * 1.1 for i in range(max_r)]

        # 4. Mesclagens (Spans)
        for merged_range in ws.merged_cells.ranges:
            s_row, s_col, e_row, e_col = merged_range.bounds
            if s_row <= max_r and s_col <= max_c:
                style_list.append(('SPAN', (s_col-1, s_row-1), (min(e_col-1, max_c-1), min(e_row-1, max_r-1))))

        # 5. Cores de Cabeçalho (Fidelidade Alisul/Supra)
        _HEADER_COLOR = colors.Color(128/255, 64/255, 0)
        headers = [5, 13, 19, 27, 32, 37, 43, 48, 54] # Linhas 6, 14, 20...
        for h_row in headers:
            style_list.append(('BACKGROUND', (0, h_row), (10, h_row), _HEADER_COLOR))
            style_list.append(('TEXTCOLOR', (0, h_row), (10, h_row), colors.white))
            style_list.append(('FONTNAME', (0, h_row), (10, h_row), 'Helvetica-Bold'))
        
        # Grade Fina
        style_list.append(('GRID', (0,5), (10, 63), 0.2, colors.grey))
        
        # Título Principal
        style_list.append(('FONTSIZE', (3,1), (10,2), 16))
        style_list.append(('ALIGN', (3,1), (10,2), 'CENTER'))

        # 6. Geração do Documento
        t = Table(matrix, colWidths=col_w, rowHeights=row_h)
        t.setStyle(TableStyle(style_list))
        tw, th = t.wrap(w - 2*margin, h)
        
        # Desenha a tabela alinhada ao topo
        t.drawOn(c, margin, h - th - 0.2*cm)

        # Inserção de Logos e QR
        if _LOGO_PATH.exists():
            c.drawImage(str(_LOGO_PATH), margin + 0.1*cm, h - 1.5*cm, width=3.8*cm, preserveAspectRatio=True, mask='auto')
        if _QR_PATH.exists():
            c.drawImage(str(_QR_PATH), w - margin - 2.5*cm, h - 1.5*cm, width=2.2*cm, preserveAspectRatio=True, mask='auto')

        c.save()
        buffer.seek(0)
        return buffer.read()

    except Exception as e:
        logger.error(f"Erro ao gerar PDF a partir do Excel: {e}", exc_info=True)
        raise RuntimeError(f"Erro técnico no PDF Supra: {str(e)}")
