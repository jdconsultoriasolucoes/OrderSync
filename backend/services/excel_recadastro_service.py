"""
backend/services/excel_recadastro_service.py
Preenche o template XLSX da Ficha Recadastro V.2026-2 utilizando estritamente as células superiores esquerdas (top-left) de cada região mesclada.
"""
import io
import os
import logging
import unicodedata
from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl.styles import Font

logger = logging.getLogger("ordersync.excel_recadastro")

TEMPLATE_PATH_DEFAULT = Path(__file__).resolve().parent.parent / "assets" / "FICHA RECADASTRO V.2026-2.xlsx"
TEMPLATE_PATH = Path(os.getenv("RECADASTRO_TEMPLATE_PATH", str(TEMPLATE_PATH_DEFAULT)))


def _br_number(value, decimals=2) -> str:
    """Formata número no padrão brasileiro (1.234,56)."""
    if value is None:
        return "0,00"
    try:
        value = float(value)
        return f"{value:,.{decimals}f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except (TypeError, ValueError):
        return "0,00"


def _s(value, default="") -> str:
    """Converte valor para string segura, tratando nulos."""
    if value is None:
        return default
    return str(value).strip()


def _normalize(text: str) -> str:
    """Remove acentos e converte para lowercase."""
    normalized = unicodedata.normalize('NFD', text)
    return ''.join(c for c in normalized if unicodedata.category(c) != 'Mn').lower().strip()


def gerar_nome_arquivo_recadastro(cliente) -> str:
    """Gera nome do arquivo: Recadastro_Cidade_NomeDoCliente."""
    cidade = _s(getattr(cliente, 'faturamento_municipio', '')) or _s(getattr(cliente, 'entrega_municipio', '')) or "SemCidade"
    nome = _s(getattr(cliente, 'cadastro_nome_cliente', '')) or "SemNome"
    cidade_safe = cidade.replace(" ", "_").replace("/", "-")
    nome_safe = nome.replace(" ", "_").replace("/", "-")
    return f"Recadastro_{cidade_safe}_{nome_safe}"


def gerar_excel_cliente_recadastro(cliente) -> bytes:
    """
    Recebe ClienteModelV2 e preenche o template da Ficha Recadastro V.2026-2.
    Respeita estritamente as células mescladas (top-left cells).
    """
    logger.info(f"Gerando Ficha Recadastro XLSX para cliente ID: {cliente.id} (Cód: {cliente.cadastro_codigo_da_empresa})")

    if not TEMPLATE_PATH.exists():
        msg = f"Template Ficha Recadastro não encontrado em: {TEMPLATE_PATH}"
        logger.error(msg)
        raise FileNotFoundError(msg)

    try:
        wb = openpyxl.load_workbook(str(TEMPLATE_PATH))
        ws = wb.active
    except Exception as e:
        msg = f"Erro ao abrir template Ficha Recadastro: {str(e)}"
        logger.error(msg, exc_info=True)
        raise RuntimeError(msg)

    try:
        # --- 1. Dados Cadastrais & Cabeçalho ---
        # A8:B8 = Rótulo 'Nome do Cliente:' | C8:H8 = Campo de Valor (Célula top-left C8)
        ws["C8"] = _s(cliente.cadastro_nome_cliente)
        
        # A9:C9 = Rótulo 'Denominação Comercial/Fantasia:' | D9:H9 = Campo de Valor (Célula top-left D9)
        ws["D9"] = _s(cliente.cadastro_nome_fantasia)
        
        # Limite Aprovado R$ (I9:K9 top-left I9)
        limite_aprovado = _br_number(cliente.cadastro_limite_credito)
        ws["I9"] = f"R $  {limite_aprovado}"

        # Contatos
        telefone = _s(getattr(cliente, 'compras_telefone_fixo_responsavel', ''))
        celular  = _s(getattr(cliente, 'compras_celular_responsavel', ''))
        email    = _s(getattr(cliente, 'compras_email_resposavel', ''))
        
        # A10:H10 = Telefone | A11:D11 = Celular | E11:K11 = E-mail
        ws["A10"] = f"Telefone:  {telefone}"
        ws["A11"] = f"Celular:  {celular}"
        ws["E11"] = f"E-mail :  {email}"

        # Documentos
        cnpj = _s(cliente.cadastro_cnpj)
        cpf  = _s(cliente.cadastro_cpf)
        ie   = _s(cliente.cadastro_inscricao_estadual)
        inscr_prod = _s(getattr(cliente, 'cadastro_inscricao_produtor', ''))

        # A12:D12 = CNPJ | E12:K12 = INSCR. PRODUTOR | A13:D13 = CPF | E13:K13 = INSCR. ESTADUAL
        ws["A12"] = f"CNPJ:  {cnpj}"
        ws["E12"] = f"INSCR. PRODUTOR:  {inscr_prod}"
        ws["A13"] = f"CPF:  {cpf}"
        ws["E13"] = f"INSCR. ESTADUAL:  {ie}"

        # --- 2. Utilização do Produto (Linha 17 - Célula A17:K17) ---
        tipo_util = _normalize(_s(cliente.cadastro_tipo_cliente))
        check_pecuaria = "    "
        check_proprio  = "    "
        check_industria = "    "
        check_revenda  = "    "

        if any(k in tipo_util for k in ["produtor rural", "pecuaria", "canil"]):
            check_pecuaria = " X  "
        elif any(k in tipo_util for k in ["pessoa fisica", "pessoa juridica", "consumidor final"]):
            check_proprio = " X  "
        elif "industrial" in tipo_util:
            check_industria = " X  "
        elif any(k in tipo_util for k in ["revenda", "lojista", "atacado"]):
            check_revenda = " X  "

        ws["A17"] = f"Utilização do Produto: ({check_pecuaria})  Insumo na Pecuária   ({check_proprio})  Consumo Próprio   ({check_industria})  Industrialização   ({check_revenda})  Comercializar/Revender"

        # --- 3. Endereço de Entrega ---
        # A20:H20 = Av/Rua/Nro | I20:K20 = CEP
        # A21:E21 = Bairro | F21:H21 = Cidade | I21:K21 = Estado
        ws["A20"] = f"Av/Rua/Nro.:  {_s(cliente.entrega_endereco)}"
        ws["I20"] = f"CEP:  {_s(cliente.entrega_cep)}"
        ws["A21"] = f"Bairro:  {_s(cliente.entrega_bairro)}"
        ws["F21"] = f"Cidade:  {_s(cliente.entrega_municipio)}"
        ws["I21"] = f"Estado:  {_s(cliente.entrega_estado)}"

        # --- 4. Endereço de Cobrança ---
        # A24:H24 = Av/Rua/Nro | I24:K24 = CEP
        # A25:E25 = Bairro | F25:H25 = Cidade | I25:K25 = Estado
        ws["A24"] = f"Av/Rua/Nro.:  {_s(cliente.cobranca_endereco)}"
        ws["I24"] = f"CEP:  {_s(cliente.cobranca_cep)}"
        ws["A25"] = f"Bairro:  {_s(cliente.cobranca_bairro)}"
        ws["F25"] = f"Cidade:  {_s(cliente.cobranca_municipio)}"
        ws["I25"] = f"Estado:  {_s(cliente.cobranca_estado)}"

        # --- 5. Limite Solicitado (A27:B27 Rótulo, C27:K27 Valor) ---
        limite_solicitado = _br_number(cliente.elaboracao_limite_credito)
        ws["C27"] = f"R$ {limite_solicitado}"

        # --- 6. Referências Bancárias (Linhas 31, 32, 33) ---
        # A31:B31 = Banco | C31:D31 = Agência | E31:F31 = Conta Corrente | G31 = Cidade | H31:I31 = Telefone | J31:K31 = Contato
        refs_bancarias = cliente.referencias_bancarias if isinstance(cliente.referencias_bancarias, list) else []
        for i, ref_b in enumerate(refs_bancarias[:3]):
            if not isinstance(ref_b, dict): continue
            row = 31 + i
            ws[f"A{row}"] = f"{i+1}- {_s(ref_b.get('banco'))}"
            ws[f"C{row}"] = _s(ref_b.get("agencia"))
            ws[f"E{row}"] = _s(ref_b.get("conta_corrente"))
            ws[f"G{row}"] = _s(ref_b.get("cidade"))
            ws[f"H{row}"] = _s(ref_b.get("telefone"))
            ws[f"J{row}"] = _s(ref_b.get("contato") or ref_b.get("gerente"))

        # --- 7. Canal de Vendas (Linhas 37, 38, 39) ---
        # A37:B37 = PET | C37:K37 = Canal Value
        ws["C37"] = _s(cliente.canal_pet)
        ws["C38"] = _s(cliente.canal_frost)
        ws["C39"] = _s(cliente.canal_insumos)

        # --- 8. Comissões DISPET (Linhas 43, 44, 45) em Vermelho ---
        _DISPET_RED = Font(color="FF0000")
        ws["C43"] = _s(cliente.comissao_pet) or "DISPET REPRESENTAÇÕES COMERCIAIS - CÓDIGO 178799"
        ws["C43"].font = _DISPET_RED

        ws["C44"] = "DISPET REPRESENTAÇÕES COMERCIAIS - CÓDIGO 178799"
        ws["C44"].font = _DISPET_RED

        ws["C45"] = _s(cliente.comissao_insumos) or "DISPET REPRESENTAÇÕES COMERCIAIS - CÓDIGO 178799"
        ws["C45"].font = _DISPET_RED

        # --- 9. Vistos / Equipe (Linhas 48, 49, 50) ---
        # E48:G48 = Nome Pet | H48:K48 = Nome Insumo
        # E49:G49 = Cod Pet  | H49:K49 = Cod Insumo
        # E50:G50 = Ger Pet  | H50:K50 = Ger Insumo
        ws["E48"] = _s(cliente.supervisor_nome_pet)
        ws["H48"] = _s(cliente.supervisor_nome_insumo)

        ws["E49"] = _s(getattr(cliente, 'supervisor_codigo_pet', ''))
        ws["H49"] = _s(getattr(cliente, 'supervisor_codigo_insumo', ''))

        ws["E50"] = _s(getattr(cliente, 'elaboracao_gerente_pet', ''))
        ws["H50"] = _s(getattr(cliente, 'elaboracao_gerente_insumos', ''))

        # --- 10. Local e Data (Linha 52) ---
        # A52:B52 Rótulo 'Local e Data' | C52:K52 Valor
        cidade_cliente = _s(cliente.faturamento_municipio) or _s(cliente.entrega_municipio)
        ws["C52"] = f"{cidade_cliente}, {datetime.now().strftime('%d/%m/%Y')}"

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()

    except Exception as e:
        logger.error(f"Erro ao gerar Ficha Recadastro XLSX para cliente {cliente.id}: {e}", exc_info=True)
        raise RuntimeError(f"Erro técnico ao processar Ficha Recadastro: {str(e)}")
