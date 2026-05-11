import openpyxl
import os

template_path = r"e:\OrderSync - Dev\backend\assets\template_supra.xlsx"

def inspect_template():
    if not os.path.exists(template_path):
        print("Template not found")
        return

    wb = openpyxl.load_workbook(template_path, data_only=True)
    ws = wb["Cadastro Parte 1"]
    print("Inspecting Cadastro Parte 1 near row 20-30:")
    for row in range(20, 31):
        line = []
        for col in range(1, 12):
            val = ws.cell(row=row, column=col).value
            line.append(f"[{openpyxl.utils.get_column_letter(col)}{row}]: {val}" if val else "---")
        print(" | ".join(line))

if __name__ == "__main__":
    inspect_template()
