import openpyxl
import os

template_path = r"e:\OrderSync - Dev\backend\assets\template_supra.xlsx"

def inspect_template():
    if not os.path.exists(template_path):
        print("Template not found")
        return

    wb = openpyxl.load_workbook(template_path, data_only=True)
    ws = wb["Cadastro Parte 1"]
    print("Inspecting Row 34 for Forma de Pagamento:")
    for col in range(1, 12):
        cell = ws.cell(row=34, column=col)
        print(f"[{openpyxl.utils.get_column_letter(col)}34]: {cell.value}")

if __name__ == "__main__":
    inspect_template()
