import openpyxl
import os

template_path = r"e:\OrderSync - Dev\backend\assets\template_supra.xlsx"

def inspect_template():
    wb = openpyxl.load_workbook(template_path)
    ws = wb["Cadastro Parte 2"]
    print("Inspecting Cadastro Parte 2, Row 10:")
    for col in range(1, 12):
        cell = ws.cell(row=10, column=col)
        print(f"[{openpyxl.utils.get_column_letter(col)}10]: {cell.value}")
    
    print("\nInspecting Cadastro Parte 2, Row 22:")
    for col in range(1, 12):
        cell = ws.cell(row=22, column=col)
        print(f"[{openpyxl.utils.get_column_letter(col)}22]: {cell.value}")

if __name__ == "__main__":
    inspect_template()
