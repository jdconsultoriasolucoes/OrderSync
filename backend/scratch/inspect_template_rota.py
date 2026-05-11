import openpyxl
import os

template_path = r"e:\OrderSync - Dev\backend\assets\template_supra.xlsx"

def inspect_template():
    if not os.path.exists(template_path):
        print("Template not found")
        return

    wb = openpyxl.load_workbook(template_path, data_only=True)
    for sheet_name in wb.sheetnames:
        print(f"\nSheet: {sheet_name}")
        ws = wb[sheet_name]
        for row in range(1, 40):
            for col in range(1, 12):
                cell = ws.cell(row=row, column=col)
                if cell.value and "rota" in str(cell.value).lower():
                    print(f"[{openpyxl.utils.get_column_letter(col)}{row}]: {cell.value}")

if __name__ == "__main__":
    inspect_template()
