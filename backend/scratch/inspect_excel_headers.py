import openpyxl

file_path = r"E:\Projeto Sistema pedidos\Planejamento\Arquivos bases edson\Entrada_Pedidos.xlsm"

def inspect_excel():
    try:
        print(f"Abrindo o arquivo: {file_path}")
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        print("Abas disponíveis:", wb.sheetnames)
        
        for sheet_name in ["Banco_Dados", "Danfes"]:
            if sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                print(f"\n--- Inspecionando Aba: {sheet_name} ---")
                
                # Pegar a primeira linha como cabeçalho
                for r in range(1, 5): # Primeiras 4 linhas
                    row_vals = [sheet.cell(row=r, column=col).value for col in range(1, 30)]
                    if any(row_vals):
                        print(f"Linha {r}: {row_vals[:26]}")
            else:
                print(f"Aba '{sheet_name}' não encontrada.")
    except Exception as e:
        print("Erro ao ler o arquivo Excel:", e)

if __name__ == "__main__":
    inspect_excel()
