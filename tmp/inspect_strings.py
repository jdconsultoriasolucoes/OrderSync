import openpyxl
from pathlib import Path

template_path = Path("e:/OrderSync/backend/assets/template_supra.xlsx")
wb = openpyxl.load_workbook(str(template_path), data_only=True)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    for row in ws.iter_rows(min_row=1, max_row=60, min_col=1, max_col=15):
        for cell in row:
            if isinstance(cell.value, str):
                val = cell.value.strip()
                if any(x in val for x in ["Vendas", "Cobrança", "C/", "P/"]):
                    print(f"[{sheet_name}] {cell.coordinate}: {repr(val)}")
