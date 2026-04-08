import openpyxl
from pathlib import Path

template_path = Path("e:/OrderSync/backend/assets/template_supra.xlsx")
if not template_path.exists():
    print(f"Template not found at {template_path}")
    exit(1)

wb = openpyxl.load_workbook(str(template_path), data_only=True)
for sheet_name in wb.sheetnames:
    print(f"\n--- Sheet: {sheet_name} ---")
    ws = wb[sheet_name]
    for row in ws.iter_rows(min_row=1, max_row=100, min_col=1, max_col=15):
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                val = cell.value.strip()
                if any(x in val for x in ["Vendas", "Cobrança", "C/", "P/"]):
                    print(f"Cell {cell.coordinate}: {repr(val)}")
                
    print("\n--- Referencias Bancarias ---")
    for row in range(32, 38):
        for col in ["A", "B", "C", "D", "E"]:
            cell = ws[f"{col}{row}"]
            if cell.value:
                print(f"Cell {col}{row}: {repr(cell.value)}")
                
    print("\n--- Referencias Comerciais ---")
    for row in range(38, 44):
        for col in ["A", "E", "G", "I"]:
            cell = ws[f"{col}{row}"]
            if cell.value:
                print(f"Cell {col}{row}: {repr(cell.value)}")
                
    print("\n--- Bens Imoveis ---")
    for row in range(44, 49):
        for col in ["A", "H", "J"]:
            cell = ws[f"{col}{row}"]
            if cell.value:
                print(f"Cell {col}{row}: {repr(cell.value)}")
                
    print("\n--- Plantel ---")
    for row in range(49, 54):
        for col in ["A", "F", "H"]:
            cell = ws[f"{col}{row}"]
            if cell.value:
                print(f"Cell {col}{row}: {repr(cell.value)}")
